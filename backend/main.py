"""
FastAPI Server for Global Business Gap Finder & Blue Ocean Intelligence Platform.
Serves both backend API and production React frontend bundle.
High-performance concurrent GeoParquet processing & Excel (.xlsx) exports.
"""

import logging
import asyncio
import os
import io
import statistics
import math
from concurrent.futures import ThreadPoolExecutor
from typing import Optional, List, Dict, Any
from fastapi import FastAPI, HTTPException, Query, Body, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from city_resolver import resolve_city_metadata
from overture_provider import fetch_city_places, resolve_latest_release
from peer_resolver import find_peer_cities
from taxonomy import TAXONOMY_MAP, CATEGORY_FAMILIES, search_categories, get_category_info, normalize_category_key
from scoring import calculate_market_gap_analysis, calculate_per_10k
from cache import (
    get_cached_city_metadata, set_cached_city_metadata,
    get_cached_city_snapshot, set_cached_city_snapshot
)

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(name)s: %(message)s")
logger = logging.getLogger("gap_finder_backend")

executor = ThreadPoolExecutor(max_workers=8)

app = FastAPI(
    title="Global Business Gap Finder API",
    description="Business Intelligence and Blue Ocean Market Opportunity Discovery API powered by Overture Maps Places & GeoParquet",
    version="1.0.0"
)

# Enable CORS for browser access
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- Input Models ---
class CityResolveRequest(BaseModel):
    country: str
    city: str

class CitySnapshotRequest(BaseModel):
    country: str
    city: str
    force_refresh: bool = False

class AnalyzeRequest(BaseModel):
    country: str
    city: str
    category_id: str
    custom_peers: Optional[List[str]] = None

class OpportunitiesRequest(BaseModel):
    country: str
    city: str
    family: Optional[str] = None
    min_market_size: Optional[int] = 0


# --- Endpoints ---

@app.get("/api/health")
async def health_check():
    release = await asyncio.to_thread(resolve_latest_release)
    return {
        "status": "online",
        "overture_release": release,
        "engine": "DuckDB GeoParquet Spatial Engine",
        "data_sources": ["Overture Maps Places", "OpenStreetMap", "Wikidata SPARQL"]
    }


@app.get("/api/countries")
async def get_countries():
    """Popular countries catalog for dropdown selector."""
    popular = [
        {"code": "BY", "name": "Belarus"},
        {"code": "GE", "name": "Georgia"},
        {"code": "ES", "name": "Spain"},
        {"code": "LT", "name": "Lithuania"},
        {"code": "DE", "name": "Germany"},
        {"code": "PL", "name": "Poland"},
        {"code": "UK", "name": "United Kingdom"},
        {"code": "FR", "name": "France"},
        {"code": "IT", "name": "Italy"},
        {"code": "GR", "name": "Greece"},
        {"code": "PT", "name": "Portugal"},
        {"code": "NL", "name": "Netherlands"},
        {"code": "BE", "name": "Belgium"},
        {"code": "AT", "name": "Austria"},
        {"code": "CH", "name": "Switzerland"},
        {"code": "SE", "name": "Sweden"},
        {"code": "NO", "name": "Norway"},
        {"code": "DK", "name": "Denmark"},
        {"code": "FI", "name": "Finland"},
        {"code": "CZ", "name": "Czech Republic"},
        {"code": "HU", "name": "Hungary"},
        {"code": "RO", "name": "Romania"},
        {"code": "IE", "name": "Ireland"},
        {"code": "UA", "name": "Ukraine"},
        {"code": "KZ", "name": "Kazakhstan"},
        {"code": "UZ", "name": "Uzbekistan"},
        {"code": "AZ", "name": "Azerbaijan"},
        {"code": "AM", "name": "Armenia"},
        {"code": "TR", "name": "Turkey"},
        {"code": "US", "name": "United States"},
        {"code": "CA", "name": "Canada"},
        {"code": "MX", "name": "Mexico"},
        {"code": "BR", "name": "Brazil"},
        {"code": "AR", "name": "Argentina"},
        {"code": "CO", "name": "Colombia"},
        {"code": "CL", "name": "Chile"},
        {"code": "PE", "name": "Peru"},
        {"code": "JP", "name": "Japan"},
        {"code": "KR", "name": "South Korea"},
        {"code": "CN", "name": "China"},
        {"code": "IN", "name": "India"},
        {"code": "TH", "name": "Thailand"},
        {"code": "VN", "name": "Vietnam"},
        {"code": "MY", "name": "Malaysia"},
        {"code": "ID", "name": "Indonesia"},
        {"code": "PH", "name": "Philippines"},
        {"code": "SG", "name": "Singapore"},
        {"code": "AU", "name": "Australia"},
        {"code": "NZ", "name": "New Zealand"},
        {"code": "EG", "name": "Egypt"},
        {"code": "NG", "name": "Nigeria"},
        {"code": "KE", "name": "Kenya"},
        {"code": "MA", "name": "Morocco"},
        {"code": "ZA", "name": "South Africa"},
        {"code": "IL", "name": "Israel"},
        {"code": "AE", "name": "United Arab Emirates"},
        {"code": "SA", "name": "Saudi Arabia"}
    ]
    return popular


@app.get("/api/cities/search")
async def search_cities(query: str = Query(..., min_length=2), country: str = Query("Georgia")):
    """Search/resolve city metadata."""
    meta = await asyncio.to_thread(resolve_city_metadata, country, query)
    if not meta:
        raise HTTPException(status_code=404, detail=f"City '{query}' in '{country}' could not be resolved.")
    return meta


@app.get("/api/categories")
async def get_categories(query: Optional[str] = None, family: Optional[str] = None):
    """Retrieve taxonomy categories and category families."""
    all_cats = search_categories(query or "")
    if family:
        all_cats = [c for c in all_cats if c.get("family") == family]
    return {
        "families": CATEGORY_FAMILIES,
        "categories": all_cats
    }


@app.post("/api/city/resolve")
async def api_resolve_city(req: CityResolveRequest):
    """Resolve city metadata, boundary, and population."""
    cached = await asyncio.to_thread(get_cached_city_metadata, req.city, req.country)
    if cached:
        return cached
    meta = await asyncio.to_thread(resolve_city_metadata, req.country, req.city)
    if not meta:
        raise HTTPException(status_code=404, detail=f"Could not resolve city '{req.city}' in '{req.country}'.")
    await asyncio.to_thread(set_cached_city_metadata, req.city, req.country, meta)
    return meta


def build_or_get_snapshot(country: str, city: str, force_refresh: bool = False) -> Dict[str, Any]:
    """Helper to load or build a cached City Snapshot."""
    if not force_refresh:
        cached = get_cached_city_snapshot(city, country)
        if cached:
            return cached
            
    meta = get_cached_city_metadata(city, country)
    if not meta:
        meta = resolve_city_metadata(country, city)
        if not meta:
            raise HTTPException(status_code=404, detail=f"Could not resolve city '{city}' in '{country}'.")
        set_cached_city_metadata(city, country, meta)
        
    bbox = meta["bbox"]
    boundary = meta.get("geojson")
    
    places = fetch_city_places(bbox, geojson_boundary=boundary)
    release = resolve_latest_release()
    
    category_counts = {}
    for p in places:
        tax_prim = p.get("taxonomy_primary") or p.get("category_primary")
        if tax_prim and tax_prim != "unclassified":
            category_counts[tax_prim] = category_counts.get(tax_prim, 0) + 1
            
    snapshot = {
        "city": meta["city"],
        "country": meta["country"],
        "display_name": meta["display_name"],
        "lat": meta["lat"],
        "lon": meta["lon"],
        "bbox": bbox,
        "geojson": boundary,
        "population": meta["population"],
        "population_year": meta["population_year"],
        "population_source": meta["population_source"],
        "fetched_at": meta.get("fetched_at", "August 2026"),
        "release": release,
        "total_places": len(places),
        "category_counts": category_counts,
        "places": places
    }
    
    set_cached_city_snapshot(city, country, snapshot)
    return snapshot


def process_single_peer(peer: dict, match_keys: set, match_hierarchy: set, category_id: str) -> Optional[dict]:
    """Process single peer city in parallel thread."""
    try:
        peer_snap = build_or_get_snapshot(peer["country"], peer["city"])
        peer_places = peer_snap["places"]
        peer_pop = peer_snap["population"]
        
        matched_peer = []
        for p in peer_places:
            tax_p = p.get("taxonomy_primary")
            cat_p = p.get("category_primary")
            bas_p = p.get("basic_category")
            hier = set(p.get("taxonomy_hierarchy", []))
            
            if (tax_p in match_keys or 
                cat_p in match_keys or 
                bas_p in match_keys or 
                bool(hier.intersection(match_hierarchy)) or
                category_id.lower() in str(tax_p).lower() or
                category_id.lower() in str(cat_p).lower()):
                matched_peer.append(p)
                
        return {
            "city": peer_snap["city"],
            "country": peer_snap["country"],
            "population": peer_pop,
            "count": len(matched_peer),
            "avg_confidence": (
                statistics.mean([p.get("confidence", 0.6) for p in matched_peer])
                if matched_peer else 0.6
            )
        }
    except Exception as e:
        logger.warning(f"Failed peer processing for {peer.get('city')}: {e}")
        return None


def run_analysis_sync(req: AnalyzeRequest) -> Dict[str, Any]:
    cat_info = get_category_info(req.category_id)
    
    target_snapshot = build_or_get_snapshot(req.country, req.city)
    target_places = target_snapshot["places"]
    target_pop = target_snapshot["population"]
    
    match_keys = set(cat_info.get("overture_keys", []))
    match_hierarchy = set(cat_info.get("hierarchy_matchers", []))
    
    matched_target_places = []
    for p in target_places:
        tax_p = p.get("taxonomy_primary")
        cat_p = p.get("category_primary")
        bas_p = p.get("basic_category")
        hier = set(p.get("taxonomy_hierarchy", []))
        
        if (tax_p in match_keys or 
            cat_p in match_keys or 
            bas_p in match_keys or 
            bool(hier.intersection(match_hierarchy)) or
            req.category_id.lower() in str(tax_p).lower() or
            req.category_id.lower() in str(cat_p).lower()):
            matched_target_places.append(p)
            
    target_count = len(matched_target_places)
    target_avg_conf = (
        statistics.mean([p.get("confidence", 0.6) for p in matched_target_places])
        if matched_target_places else 0.6
    )
    
    peer_candidates = find_peer_cities(target_snapshot["city"], target_snapshot["country"], target_pop, limit=5)
    
    futures = [
        executor.submit(process_single_peer, peer, match_keys, match_hierarchy, req.category_id)
        for peer in peer_candidates
    ]
    
    peer_analysis_data = []
    for f in futures:
        res = f.result()
        if res:
            peer_analysis_data.append(res)
            
    analysis_res = calculate_market_gap_analysis(
        target_city=target_snapshot["city"],
        target_population=target_pop,
        target_existing_count=target_count,
        target_avg_poi_confidence=target_avg_conf,
        category_title=cat_info["title"],
        peer_data=peer_analysis_data,
        population_year=target_snapshot["population_year"]
    )
    
    bbox = target_snapshot["bbox"]
    mid_x = (bbox[0] + bbox[2]) / 2.0
    mid_y = (bbox[1] + bbox[3]) / 2.0
    
    quadrants = {
        "North-West": 0,
        "North-East": 0,
        "South-West": 0,
        "South-East": 0
    }
    
    for p in matched_target_places:
        lon, lat = p["lon"], p["lat"]
        if lat >= mid_y and lon < mid_x:
            quadrants["North-West"] += 1
        elif lat >= mid_y and lon >= mid_x:
            quadrants["North-East"] += 1
        elif lat < mid_y and lon < mid_x:
            quadrants["South-West"] += 1
        else:
            quadrants["South-East"] += 1
            
    analysis_res["category_info"] = cat_info
    analysis_res["matched_places"] = matched_target_places
    analysis_res["city_metadata"] = {
        "city": target_snapshot["city"],
        "country": target_snapshot["country"],
        "lat": target_snapshot["lat"],
        "lon": target_snapshot["lon"],
        "bbox": target_snapshot["bbox"],
        "geojson": target_snapshot["geojson"],
        "population": target_pop,
        "population_year": target_snapshot["population_year"],
        "population_source": target_snapshot["population_source"],
        "release": target_snapshot["release"]
    }
    analysis_res["neighborhood_density"] = quadrants
    return analysis_res


@app.post("/api/city/snapshot")
async def api_city_snapshot(req: CitySnapshotRequest):
    return await asyncio.to_thread(build_or_get_snapshot, req.country, req.city, req.force_refresh)


@app.post("/api/analyze")
async def api_analyze_market(req: AnalyzeRequest):
    """MODE A — Complete Market Gap Analysis for a specific Industry in a City."""
    return await asyncio.to_thread(run_analysis_sync, req)


def run_opportunities_sync(req: OpportunitiesRequest) -> Dict[str, Any]:
    target_snapshot = build_or_get_snapshot(req.country, req.city)
    target_places = target_snapshot["places"]
    target_pop = target_snapshot["population"]
    
    peer_candidates = find_peer_cities(target_snapshot["city"], target_snapshot["country"], target_pop, limit=4)
    
    def fetch_peer_snap(peer):
        try:
            return build_or_get_snapshot(peer["country"], peer["city"])
        except Exception as e:
            logger.warning(f"Error loading peer {peer.get('city')}: {e}")
            return None
            
    peer_futures = [executor.submit(fetch_peer_snap, p) for p in peer_candidates]
    peer_snapshots = [f.result() for f in peer_futures if f.result() is not None]
            
    opportunities = []
    
    for cat_id, cat_info in TAXONOMY_MAP.items():
        if req.family and cat_info["family"] != req.family:
            continue
            
        match_keys = set(cat_info.get("overture_keys", []))
        match_hierarchy = set(cat_info.get("hierarchy_matchers", []))
        
        target_count = sum(
            1 for p in target_places
            if p.get("taxonomy_primary") in match_keys or 
               p.get("category_primary") in match_keys or 
               p.get("basic_category") in match_keys or
               bool(set(p.get("taxonomy_hierarchy", [])).intersection(match_hierarchy))
        )
        
        peer_data = []
        for p_snap in peer_snapshots:
            p_count = sum(
                1 for p in p_snap["places"]
                if p.get("taxonomy_primary") in match_keys or 
                   p.get("category_primary") in match_keys or 
                   p.get("basic_category") in match_keys or
                   bool(set(p.get("taxonomy_hierarchy", [])).intersection(match_hierarchy))
            )
            peer_data.append({
                "city": p_snap["city"],
                "country": p_snap["country"],
                "population": p_snap["population"],
                "count": p_count,
                "avg_confidence": 0.7
            })
            
        analysis = calculate_market_gap_analysis(
            target_city=target_snapshot["city"],
            target_population=target_pop,
            target_existing_count=target_count,
            target_avg_poi_confidence=0.7,
            category_title=cat_info["title"],
            peer_data=peer_data,
            population_year=target_snapshot["population_year"]
        )
        
        opp_obj = {
            "category_id": cat_id,
            "category_title": cat_info["title"],
            "family": cat_info["family"],
            "family_title": CATEGORY_FAMILIES.get(cat_info["family"], {}).get("title", "Services"),
            "existing_count": target_count,
            "per_10k": analysis["per_10k"],
            "benchmark_per_10k": analysis["benchmark_per_10k"],
            "expected_count": analysis["expected_count"],
            "estimated_gap": analysis["estimated_gap"],
            "gap_percent": analysis["gap_percent"],
            "opportunity_score": analysis["opportunity_score"],
            "opportunity_label": analysis["opportunity_label"],
            "data_confidence_score": analysis["data_confidence_score"]
        }
        opportunities.append(opp_obj)
        
    opportunities.sort(key=lambda x: x["opportunity_score"], reverse=True)
    
    return {
        "city": target_snapshot["city"],
        "country": target_snapshot["country"],
        "population": target_pop,
        "population_year": target_snapshot["population_year"],
        "total_categories_scanned": len(opportunities),
        "opportunities": opportunities
    }


@app.post("/api/opportunities")
async def api_discover_opportunities(req: OpportunitiesRequest):
    """MODE B — Discover Opportunities (Gap Scanner across all taxonomy categories)."""
    return await asyncio.to_thread(run_opportunities_sync, req)


@app.post("/api/export/excel")
async def export_excel(req: AnalyzeRequest):
    """Generate and stream rich Excel (.xlsx) workbook for market analysis."""
    analysis = await asyncio.to_thread(run_analysis_sync, req)
    
    wb = openpyxl.Workbook()
    
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    title_font = Font(name="Calibri", size=14, bold=True, color="0F172A")
    subtitle_font = Font(name="Calibri", size=10, italic=True, color="64748B")
    bold_font = Font(name="Calibri", size=11, bold=True)
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # --- Sheet 1: Executive Summary ---
    ws1 = wb.active
    ws1.title = "Executive Summary"
    
    ws1.cell(row=1, column=1, value=f"Market Intelligence Report: {analysis['category_title']} in {analysis['target_city']}, {req.country}").font = title_font
    ws1.cell(row=2, column=1, value=f"Population: {analysis['target_population']:,} ({analysis['population_year']}) | Generated via GapFinder.ai").font = subtitle_font
    
    metrics = [
        ("Target City", analysis["target_city"]),
        ("Country", req.country),
        ("Industry", analysis["category_title"]),
        ("Population", f"{analysis['target_population']:,}"),
        ("Existing Business Count", analysis["existing_count"]),
        ("Businesses per 10,000 Residents", f"{analysis['per_10k']:.2f}"),
        ("Peer City Benchmark per 10k", f"{analysis['benchmark_per_10k']:.2f}"),
        ("Expected Supply at Benchmark", analysis["expected_count"]),
        ("Estimated Supply Gap", f"{'+' if analysis['estimated_gap'] > 0 else ''}{analysis['estimated_gap']}"),
        ("Deficit / Surplus Percentage", f"{analysis['gap_percent']}%"),
        ("Opportunity Score (0-100)", f"{analysis['opportunity_score']}/100 ({analysis['opportunity_label']})"),
        ("Data Confidence Score (0-100)", f"{analysis['data_confidence_score']}/100"),
        ("Analysis Date", "August 2026"),
        ("Overture Maps Release", analysis['city_metadata']['release'])
    ]
    
    ws1.cell(row=4, column=1, value="Metric").font = header_font
    ws1.cell(row=4, column=1).fill = header_fill
    ws1.cell(row=4, column=2, value="Value").font = header_font
    ws1.cell(row=4, column=2).fill = header_fill
    
    for r_idx, (k, v) in enumerate(metrics, start=5):
        c1 = ws1.cell(row=r_idx, column=1, value=k)
        c2 = ws1.cell(row=r_idx, column=2, value=str(v))
        c1.font = bold_font
        c1.border = thin_border
        c2.border = thin_border
        
    ws1.cell(row=20, column=1, value="Executive Narrative Summary:").font = bold_font
    ws1.cell(row=21, column=1, value=analysis["explanation"])
    
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 50
    
    # --- Sheet 2: Businesses Directory ---
    ws2 = wb.create_sheet(title="Detected Establishments")
    
    headers = [
        "Business Name", "Primary Category", "Basic Category", "Address / Locality", 
        "Brand", "Phone", "Email", "Website", "Google Maps Profile URL", 
        "Confidence Score", "Operating Status", "Longitude", "Latitude"
    ]
    ws2.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        
    for p in analysis["matched_places"]:
        google_maps_url = f"https://www.google.com/maps/search/?api=1&query={p.get('name', '')} {p.get('address', '') or p.get('locality', '')}"
        row_vals = [
            p.get("name", "Unnamed"),
            p.get("taxonomy_primary", p.get("category_primary")),
            p.get("basic_category"),
            p.get("address") or p.get("locality") or "N/A",
            p.get("brand") or "N/A",
            p.get("phone") or "N/A",
            p.get("email") or "N/A",
            p.get("website") or "N/A",
            google_maps_url,
            f"{p.get('confidence', 0.5) * 100:.0f}%",
            p.get("operating_status", "operating"),
            p.get("lon"),
            p.get("lat")
        ]
        ws2.append(row_vals)
        
    for col_idx in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 22
        
    # --- Sheet 3: Peer Cities Benchmarks ---
    ws3 = wb.create_sheet(title="Peer Cities Benchmarks")
    peer_headers = ["Peer City", "Country", "Population", "Existing POI Count", "Businesses per 10k Residents"]
    ws3.append(peer_headers)
    for col_idx in range(1, len(peer_headers) + 1):
        cell = ws3.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        
    for peer in analysis["peer_cities"]:
        ws3.append([
            peer.get("city"),
            peer.get("country"),
            peer.get("population"),
            peer.get("existing_count"),
            peer.get("per_10k")
        ])
    for col_idx in range(1, len(peer_headers) + 1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = 25

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    filename = f"{analysis['target_city']}_{req.category_id}_analysis.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# Mount static frontend production build if directory exists
dist_dir = os.path.join(os.path.dirname(__file__), "../frontend/dist")
if os.path.exists(dist_dir):
    app.mount("/assets", StaticFiles(directory=os.path.join(dist_dir, "assets")), name="assets")

    @app.get("/")
    async def serve_index():
        index_file = os.path.join(dist_dir, "index.html")
        if os.path.exists(index_file):
            return FileResponse(index_file)
        raise HTTPException(status_code=404, detail="Index file not found")

    @app.get("/{full_path:path}")
    async def serve_frontend(full_path: str):
        if full_path.startswith("api"):
            raise HTTPException(status_code=404, detail="API route not found")
        index_file = os.path.join(dist_dir, "index.html")
        if os.path.exists(index_file):
            return FileResponse(index_file)
        raise HTTPException(status_code=404, detail="Index file not found")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=3000, reload=False)
